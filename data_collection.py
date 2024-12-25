from pylsl import StreamInlet, resolve_stream
import numpy as np
import time
from scipy.signal import welch, butter, filtfilt
import matplotlib.pyplot as plt
from collections import deque
import pandas as pd
import random
from openpyxl import load_workbook,  Workbook

import os
import random
import string

#Time
start_time = time.time()

# Parameters
fs = 256 #Adjust based on device
buffer_length = 0.5  # Reduced buffer length to 0.5 seconds
num_samples = int(fs * buffer_length)  # Total number of samples to collect per buffer
max_points = 200  # Increased number of points to display in the plot, more points results in smoother visualization



#Data Storage:
times = deque(maxlen=max_points)

delta_powers_tp9 = deque(maxlen=max_points) #double ended queue. Automatically removes older points when new ones append in a list of size = max_points
delta_powers_af7 = deque(maxlen=max_points)
delta_powers_af8 = deque(maxlen=max_points)
delta_powers_tp10 = deque(maxlen=max_points)

alpha_powers_tp9 = deque(maxlen=max_points)
alpha_powers_af7 = deque(maxlen=max_points)
alpha_powers_af8 = deque(maxlen=max_points)
alpha_powers_tp10 = deque(maxlen=max_points)

beta_powers_tp9 = deque(maxlen=max_points)
beta_powers_af7 = deque(maxlen=max_points)
beta_powers_af8 = deque(maxlen=max_points)
beta_powers_tp10 = deque(maxlen=max_points)

gamma_powers_tp9 = deque(maxlen=max_points)
gamma_powers_af7 = deque(maxlen=max_points)
gamma_powers_af8 = deque(maxlen=max_points)
gamma_powers_tp10 = deque(maxlen=max_points)

theta_powers_tp9 = deque(maxlen=max_points)
theta_powers_af7 = deque(maxlen=max_points)
theta_powers_af8 = deque(maxlen=max_points)
theta_powers_tp10 = deque(maxlen=max_points)



# Connect to Device:
print("Looking for an EEG stream...")
streams = resolve_stream('type', 'EEG')
if not streams:
    print("No EEG stream found.")
    exit()
inlet = StreamInlet(streams[0]) #Create data transfer Stream/Pipeline
print("Starting data acquisition...")

tp9_buffer = [] #Buffer for temorary data storage
af7_buffer = []
af8_buffer = []
tp10_buffer = []



# Plotting
plt.ion()
fig, ax = plt.subplots()
line_tp9, = ax.plot([], [], label='TP9')
line_af7, = ax.plot([], [], label='AF7')
line_af8, = ax.plot([], [], label='AF8')
line_tp10, = ax.plot([], [], label='TP10')
ax.set_xlabel('Time (s)')
ax.set_ylabel('Gamma Band Power (µV²)')
ax.set_title('Gamma Band Power Over Time')
ax.legend()
plt.show()



# *** Data simulation:
def simulate_inlet_pull_sample():
    # Generate a sample with 4 random values
    sample = [random.uniform(-100, 100) for _ in range(4)]
    timestamp = time.time()
    return sample, timestamp



#Appending data:
def create_excel_if_not_exists(file_path):
    if not os.path.exists(file_path):
        try:
            wb = Workbook()
            wb.save(file_path)
            print(f"Excel file created at: {file_path}")
            return True
        except Exception as e:
            print(f"An error occurred while creating the Excel file: {e}")
            return False
    else:
        print(f"Excel file already exists at: {file_path}")
        return False
    
def store_deque_in_excel(deq, row_num, excel_file):
    create_excel_if_not_exists(excel_file)

    # Validate inputs
    if not isinstance(deq, deque):
        raise TypeError("The first argument must be a deque.")
    if not isinstance(row_num, int) or row_num <= 0:
        raise ValueError("Row number must be a positive integer.")
    if not os.path.isfile(excel_file):
        raise IOError(f"The file {excel_file} does not exist.")

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        for col_num, value in enumerate(deq, start=1):
            ws.cell(row=row_num, column=col_num, value=value)
        wb.save(excel_file)
    except Exception as e:
        raise IOError(f"An error occurred while processing the Excel file: {e}")
    


#Average:
def average_first_n_values(data_deque, n):
    data_list = list(data_deque)  # Convert deque to list
    return sum(data_list[:n]) / n if len(data_list) >= n else sum(data_list) / len(data_list)



#Ahthorization:
print("Press Enter to start Data collection: ")
data = input()
print("\n COLLECTING...")



#Collect & Process Data:
try:
    while True:
        # Collect raw/noisy data in buffer
        for _ in range(num_samples):
            sample, timestamp = inlet.pull_sample()
            # sample, timestamp = simulate_inlet_pull_sample()
            tp9_buffer.append(sample[0])
            af7_buffer.append(sample[1])
            af8_buffer.append(sample[2])
            tp10_buffer.append(sample[3])

        tp9_data = np.array(tp9_buffer) #Tranform into numpy array for easier operations 
        af7_data = np.array(af7_buffer)
        af8_data = np.array(af8_buffer)
        tp10_data = np.array(tp10_buffer)





        # Bandpass filter to remove noise outside EEG frequency range
        def bandpass_filter(data, lowcut, highcut, fs, order=4): #lowcut = lowest frequency to be kept; highcut = highest frequency to be kept.
            nyquist = 0.5 * fs  #nyquest frequency, half of sampling frequency. Based on Nyquists theorem: it's the highest freuqncy that can be accurately represented at a given sampling rate.
            low = lowcut / nyquist  #Normalizing low and high using Norquest frequency
            high = highcut / nyquist
            b, a = butter(order, [low, high], btype='band') #Butterworth filter, get parameters for that filter (?)
            y = filtfilt(b, a, data)    #Applies butterworth filter to data
            return y  #return data after filtering
        lowcut = 0.5
        highcut = 50.0
        tp9_data = bandpass_filter(tp9_data, lowcut, highcut, fs)
        af7_data = bandpass_filter(af7_data, lowcut, highcut, fs)
        af8_data = bandpass_filter(af8_data, lowcut, highcut, fs)
        tp10_data = bandpass_filter(tp10_data, lowcut, highcut, fs)





        # Compute band powers (power of the wave)
        def compute_band_powers(data, fs):
            freqs, psd = welch(data, fs, nperseg=len(data)) # Compute the Power Spectral Density (PSD): freqs = array of frequency bins. psd = coresponding power value of each bin
            
            bands = { # Define frequency bands mapping frequency to band names
                'Delta': (0.5, 4),
                'Theta': (4, 8),
                'Alpha': (8, 13),
                'Beta': (13, 30),
                'Gamma': (30, 50)
            }

            band_powers = {}
            for band_name, (low, high) in bands.items():    #iterate over the bands dictionary
                # Find indices corresponding to the frequency band
                idx_band = np.logical_and(freqs >= low, freqs <= high) #if frequency is within the stated range, return TRUE. idx_band = boolean array
                # Integrate the PSD over the frequency band
                band_power = np.trapz(psd[idx_band], freqs[idx_band]) #calculate definite integral (area under curve) which represents power
                band_powers[band_name] = band_power #
            return band_powers  #return dictionary, with each power values for each band
        
        tp9_band_powers = compute_band_powers(tp9_data, fs)
        af7_band_powers = compute_band_powers(af7_data, fs)
        af8_band_powers = compute_band_powers(af8_data, fs)
        tp10_band_powers = compute_band_powers(tp10_data, fs)





        # Data storage
        delta_powers_tp9.append(tp9_band_powers['Delta']) # Store Delta band power
        delta_powers_af7.append(af7_band_powers['Delta'])
        delta_powers_af8.append(af8_band_powers['Delta'])
        delta_powers_tp10.append(tp10_band_powers['Delta'])     #High mucle movement = Higher tp10 delta

        alpha_powers_tp9.append(tp9_band_powers['Alpha'])
        alpha_powers_af7.append(af7_band_powers['Alpha'])
        alpha_powers_af8.append(af8_band_powers['Alpha'])
        alpha_powers_tp10.append(tp10_band_powers['Alpha'])

        beta_powers_tp9.append(tp9_band_powers['Beta'])
        beta_powers_af7.append(af7_band_powers['Beta'])
        beta_powers_af8.append(af8_band_powers['Beta'])
        beta_powers_tp10.append(tp10_band_powers['Beta'])

        gamma_powers_tp9.append(tp9_band_powers['Gamma'])
        gamma_powers_af7.append(af7_band_powers['Gamma'])
        gamma_powers_af8.append(af8_band_powers['Gamma'])
        gamma_powers_tp10.append(tp10_band_powers['Gamma'])

        theta_powers_tp9.append(tp9_band_powers['Theta'])
        theta_powers_af7.append(af7_band_powers['Theta'])
        theta_powers_af8.append(af8_band_powers['Theta'])
        theta_powers_tp10.append(tp10_band_powers['Theta'])

        current_time = time.time() - start_time
        times.append(current_time)

        # Updating Plot
        line_tp9.set_data(times, gamma_powers_tp9)
        line_af7.set_data(times, gamma_powers_af7)
        line_af8.set_data(times, gamma_powers_af8)
        line_tp10.set_data(times, gamma_powers_tp10)
        ax.relim()
        ax.autoscale_view()
        plt.pause(0.001)  # Decreased pause time for faster updates

        # Clear buffers for next point
        tp9_buffer.clear()
        af7_buffer.clear()
        af8_buffer.clear()
        tp10_buffer.clear()




except KeyboardInterrupt:
    print("Appending Data, Please Wait...")
    
    #Average:
    values_to_average = 15
    average_delta_tp9 = average_first_n_values(delta_powers_tp9, values_to_average)
    average_delta_af7 = average_first_n_values(delta_powers_af7, values_to_average)
    average_delta_af8 = average_first_n_values(delta_powers_af8, values_to_average)
    average_delta_tp10 = average_first_n_values(delta_powers_tp10, values_to_average)

    average_alpha_tp9 = average_first_n_values(alpha_powers_tp9, values_to_average)
    average_alpha_af7 = average_first_n_values(alpha_powers_af7, values_to_average)
    average_alpha_af8 = average_first_n_values(alpha_powers_af8, values_to_average)
    average_alpha_tp10 = average_first_n_values(alpha_powers_tp10, values_to_average)

    average_beta_tp9 = average_first_n_values(beta_powers_tp9, values_to_average)
    average_beta_af7 = average_first_n_values(beta_powers_af7, values_to_average)
    average_beta_af8 = average_first_n_values(beta_powers_af8, values_to_average)
    average_beta_tp10 = average_first_n_values(beta_powers_tp10, values_to_average)

    average_gamma_tp9 = average_first_n_values(gamma_powers_tp9, values_to_average)
    average_gamma_af7 = average_first_n_values(gamma_powers_af7, values_to_average)
    average_gamma_af8 = average_first_n_values(gamma_powers_af8, values_to_average)
    average_gamma_tp10 = average_first_n_values(gamma_powers_tp10, values_to_average)

    average_theta_tp9 = average_first_n_values(theta_powers_tp9, values_to_average)
    average_theta_af7 = average_first_n_values(theta_powers_af7, values_to_average)
    average_theta_af8 = average_first_n_values(theta_powers_af8, values_to_average)
    average_theta_tp10 = average_first_n_values(theta_powers_tp10, values_to_average)

    dq1 = deque([average_delta_tp9, average_theta_tp9, average_alpha_tp9, average_beta_tp9, average_gamma_tp9])  # TP9
    dq2 = deque([average_delta_af7, average_theta_af7, average_alpha_af7, average_beta_af7, average_gamma_af7])  # AF7
    dq3 = deque([average_delta_af8, average_theta_af8, average_alpha_af8, average_beta_af8, average_gamma_af8])  # AF8
    dq4 = deque([average_delta_tp10, average_theta_tp10, average_alpha_tp10, average_beta_tp10, average_gamma_tp10])  # TP10

    store_deque_in_excel(dq1, row_num=1, excel_file="data.xlsx" )
    store_deque_in_excel(dq2, row_num=2, excel_file="data.xlsx" )
    store_deque_in_excel(dq3, row_num=3, excel_file="data.xlsx" )
    store_deque_in_excel(dq4, row_num=4, excel_file="data.xlsx" )

    data = delta_powers_tp9
    store_deque_in_excel(data, row_num=1, excel_file="data/data_entirety_delta.xlsx" )
    data = delta_powers_af7
    store_deque_in_excel(data, row_num=2, excel_file="data/data_entirety_delta.xlsx" )
    data = delta_powers_af8
    store_deque_in_excel(data,row_num=3, excel_file="data/data_entirety_delta.xlsx" )
    data = delta_powers_tp10
    store_deque_in_excel(data, row_num=4, excel_file="data/data_entirety_delta.xlsx" )

    data = alpha_powers_tp9
    store_deque_in_excel(data, row_num=1, excel_file="./data/data_entirety_alpha.xlsx" )
    data = delta_powers_af7
    store_deque_in_excel(data, row_num=2, excel_file="./data/data_entirety_alpha.xlsx" )
    data = delta_powers_af8
    store_deque_in_excel(data,row_num=3, excel_file="./data/data_entirety_alpha.xlsx")
    data = delta_powers_tp10
    store_deque_in_excel(data, row_num=4, excel_file="./data/data_entirety_alpha.xlsx")

    data = beta_powers_tp9
    store_deque_in_excel(data, row_num=1, excel_file="./data/data_entirety_beta.xlsx" )
    data = beta_powers_af7
    store_deque_in_excel(data, row_num=2, excel_file="./data/data_entirety_beta.xlsx" )
    data = beta_powers_af8
    store_deque_in_excel(data,row_num=3, excel_file="./data/data_entirety_beta.xlsx")
    data = beta_powers_tp10
    store_deque_in_excel(data, row_num=4, excel_file="./data/data_entirety_beta.xlsx")

    data = gamma_powers_tp9
    store_deque_in_excel(data, row_num=1, excel_file="./data/data_entirety_gamma.xlsx" )
    data = gamma_powers_af7
    store_deque_in_excel(data, row_num=2, excel_file="./data/data_entirety_gamma.xlsx" )
    data = gamma_powers_af8
    store_deque_in_excel(data,row_num=3, excel_file="./data/data_entirety_gamma.xlsx" )
    data = gamma_powers_tp10
    store_deque_in_excel(data, row_num=4, excel_file="./data/data_entirety_gamma.xlsx" )

    data = theta_powers_tp9
    store_deque_in_excel(data, row_num=1, excel_file="./data/data_entirety_theta.xlsx" )
    data = beta_powers_af7
    store_deque_in_excel(data, row_num=2, excel_file="./data/data_entirety_theta.xlsx" )
    data = beta_powers_af8
    store_deque_in_excel(data,row_num=3, excel_file="./data/data_entirety_theta.xlsx")
    data = beta_powers_tp10
    store_deque_in_excel(data, row_num=4, excel_file="./data/data_entirety_theta.xlsx")

    print("Data acquisition stopped.")